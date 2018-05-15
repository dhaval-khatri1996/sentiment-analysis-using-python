import openpyxl 
def getDataset():
	wb = openpyxl.load_workbook('dataset.xlsx')
	sheet = wb['Sheet1']
	arr = {}
	#Read data from row 1 to the last row
	for i in range(1 , sheet.max_row):
				key = sheet.cell(row=i, column=1).value
				polarity = sheet.cell(row=i, column=2).value
				if(key!= None):
					#the word is the key and the value is polarity
					arr[key.lower()]=polarity
	return arr