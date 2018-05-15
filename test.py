
import nltk
import openpyxl
import re

from treeGenerator import generateTree
from createDictionary import getDataset
import word 

def specialCharacter(sentence):
    sentence = re.sub('(?<=\w)([!?,.:;/\\<>"])', r' \1', sentence)
    return sentence.lower().split()

tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')

#Get the list of words and their polarity  
wordDataset = getDataset()


wb = openpyxl.load_workbook('test.xlsx')
sheet = wb['Sheet1']
correctCount = 0
for j in range(1 , sheet.max_row+1):
    sentences = sheet.cell(row=j,column=1).value
    polarity = sheet.cell(row=j,column=2).value
    if(sentences=='' or sentences==None or polarity == None):
        continue 
    sentences = tokenizer.tokenize(sentences)
    avg =0
    count =0
    for sentence in sentences:
        #Split the sentence into word
        wordList = nltk.pos_tag(specialCharacter(sentence),tagset='universal' )
        words=[]
    
        for i  in wordList:
            if ( i[0] in wordDataset):
                words.append(word.Word(i[0],wordDataset[i[0]],i[1]))
            else:
                words.append(word.Word(i[0],0,"X"))
        ans = generateTree(words)
        if(ans>10):
            ans=10
        elif(ans<-10):
            ans=-10
        avg +=ans
        count=count+1
    #print('{},{}'.format(j, avg ))
    sheet.cell(row=j, column=3, value= 0 if avg<=0 else 1)
    if((polarity==1 and avg>0)or (polarity==0 and avg<=0)):
        correctCount +=1
print("Out of {}, {} reviews were rated correctly \n {} accuracy"
        .format(sheet.max_row+1,correctCount, round((correctCount/(sheet.max_row+1))*100,2)))
wb.save("dumy1.xlsx")